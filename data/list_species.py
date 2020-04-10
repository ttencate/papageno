#!/usr/bin/env python

'''
Traverses the official IOC World Bird List
(https://www.worldbirdnames.org/ioc-lists/master-list-2/), the multilingual
version thereof, to produce a machine-readable list of all bird species.

Also assigns stable, unique species ids.
'''

import argparse
import csv
import logging
import os
import re
import sys

import openpyxl


_LANGUAGE_CODES = {
    'English': 'en',
    'Afrikaans': 'af',
    'Catalan': 'ca',
    # https://stackoverflow.com/questions/4892372/language-codes-for-simplified-chinese-and-traditional-chinese
    'Chinese': 'zh_CN',
    'Chinese (Traditional)': 'zh_TW',
    'Czech': 'ics',
    'Danish': 'da',
    'Dutch': 'nl',
    'Estonian': 'et',
    'Finnish': 'fi',
    'French': 'fr',
    'German': 'de',
    'Hungarian': 'hu',
    'Icelandic': 'is',
    'Indonesian': 'id',
    'Italian': 'it',
    'Japanese': 'ja',
    'Latvian': 'lv',
    'Lithuanian': 'lt',
    'Northern Sami': 'se',
    'Norwegian': 'no',
    'Polish': 'pl',
    'Portuguese': 'pt',
    'Russian': 'ru',
    'Slovak': 'sk',
    'Slovenian': 'sl',
    'Spanish': 'es',
    'Swedish': 'sv',
    'Thai': 'th',
    'Ukrаiniаn': 'uk',
}


class Multiling:

    _NUM_HEADER_ROWS = 3

    def __init__(self, file_name):
        workbook = openpyxl.load_workbook(file_name, read_only=True)
        self._worksheet = workbook['List']
        self.fields = []
        for row in self._worksheet.iter_rows(min_row=1, max_row=Multiling._NUM_HEADER_ROWS, values_only=True):
            if not self.fields:
                self.fields = [None] * len(row)
            for i, cell in enumerate(row):
                if cell and not self.fields[i]:
                    self.fields[i] = cell

    def iter_species(self):
        restart_column = self.fields.index('Scientific Name')
        merged_row = None
        for row in self._worksheet.iter_rows(min_row=Multiling._NUM_HEADER_ROWS + 1, values_only=True):
            if row[restart_column]:
                if merged_row and merged_row['Scientific Name']:
                    yield merged_row
                merged_row = {field: None for field in self.fields}
            if merged_row:
                for field, value in zip(self.fields, row):
                    if value:
                        merged_row[field] = value
        yield merged_row


class SpeciesList:

    def __init__(self, file_name=None):
        self._rows = []
        self._indices_by_scientific_name = {}
        if file_name:
            with open(file_name, 'rt') as input_file:
                reader = csv.DictReader(input_file)
                self.fields = reader.fieldnames
                for row in reader:
                    self._indices_by_scientific_name[row['scientific_name']] = len(self._rows)
                    self._rows.append(row)
        else:
            self.fields = ['species_id', 'scientific_name']
        self._next_species_id = max(row['species_id'] for row in self._rows) if self._rows else 1

    def __len__(self):
        return len(self._rows)

    def add_column(self, language_code):
        self.fields.append(language_code)
        for row in self._rows:
            row[language_code] = ''

    def get_species(self, scientific_name):
        return self._rows[self._indices_by_scientific_name[scientific_name]]
    
    def add_species(self, scientific_name):
        row = {key: None for key in self.fields}
        row['species_id'] = self._next_species_id
        self._next_species_id += 1
        row['scientific_name'] = scientific_name
        self._indices_by_scientific_name[scientific_name] = len(self._rows)
        self._rows.append(row)
        return row

    def write(self, file_name):
        with open(file_name, 'wt') as output_file:
            writer = csv.DictWriter(output_file, self.fields)
            writer.writeheader()
            for row in self._rows:
                writer.writerow(row)


def _main():
    logging.basicConfig(level=logging.INFO)

    parser = argparse.ArgumentParser()
    parser.add_argument(
        '--ioc_multiling_file',
        default=os.path.join(os.path.dirname(__file__), 'sources', 'Multiling IOC 9.1b.xlsx'), 
        help='Path to the multilingual Excel file downloaded from IOC')
    parser.add_argument(
        '--species_list_file',
        default=os.path.join(os.path.dirname(__file__), 'sources', 'species.csv'), 
        help='File of species list to be created/updated')
    args = parser.parse_args()

    logging.info(f'Loading {args.ioc_multiling_file}')
    multiling = Multiling(args.ioc_multiling_file)
    logging.info(f'Found column headings: {multiling.fields}')

    try:
        species_list = SpeciesList(args.species_list_file)
        logging.info(f'Read {args.species_list_file} containing {len(species_list)} species')
    except FileNotFoundError:
        logging.info(f'Could not find {args.species_list_file}, starting afresh')
        species_list = SpeciesList()

    for field in multiling.fields:
        language_code = _LANGUAGE_CODES.get(field)
        if not language_code:
            if not re.match(r'^(No\d.*|Order|Family|Scientific Name)$', field):
                raise ValueError(f'Do not know what to do with column {field}')
            continue
        if language_code not in species_list.fields:
            logging.info(f'Adding new output column {language_code}')
            species_list.add_column(language_code)

    for row in multiling.iter_species():
        scientific_name = row['Scientific Name']
        try:
            dest_row = species_list.get_species(scientific_name)
        except KeyError:
            dest_row = species_list.add_species(scientific_name)
            logging.info(f'Added new species {scientific_name} (id {dest_row["species_id"]})')
        for field, value in row.items():
            language_code = _LANGUAGE_CODES.get(field)
            if language_code:
                dest_row[language_code] = value

    logging.info(f'Writing output to {args.species_list_file}')
    species_list.write(args.species_list_file)
    

if __name__ == '__main__':
    sys.exit(_main())
