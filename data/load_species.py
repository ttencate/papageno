'''
Parses the official [IOC World Bird
List](https://www.worldbirdnames.org/ioc-lists/master-list-2/) spreadsheet
(Multilingual Version). It outputs to the `species` table and the
`common_names` table.

Note that what constitutes a "species" changes as scientific insight
progresses, which is why the IOC releases new lists every once in a while. For
our purposes, species = scientific name.

xeno-canto also uses the sheets from IOC as their source; see [the Articles
section](https://www.xeno-canto.org/articles) on the site for updates about
which version they last updated to. For best results, we should use the same
version.
'''

import logging
import os
import re

import openpyxl
from sqlalchemy.orm import joinedload

from species import Species, CommonName


_LANGUAGE_MAPPING = {
    'English': 'en',
    'Afrikaans': 'af',
    'Catalan': 'ca',
    # https://stackoverflow.com/questions/4892372/language-codes-for-simplified-chinese-and-traditional-chinese
    'Chinese': 'zh_CN',
    'Chinese (Traditional)': 'zh_TW',
    'Czech': 'cs',
    'Danish': 'da',
    'Dutch': 'nl',
    'Estonian': 'et',
    'Finnish': 'fi',
    'French': 'fr',
    'German': 'de',
    'Hungarian': 'hu',
    'Icelandic': 'is',
    'Indonesian': 'id',
    'Italian': 'it',
    'Japanese': 'ja',
    'Latvian': 'lv',
    'Lithuanian': 'lt',
    'Northern Sami': 'se',
    'Norwegian': 'no',
    'Polish': 'pl',
    'Portuguese': 'pt',
    'Russian': 'ru',
    'Slovak': 'sk',
    'Slovenian': 'sl',
    'Spanish': 'es',
    'Swedish': 'sv',
    'Thai': 'th',
    'Ukrаiniаn': 'uk',
}


class Multiling:
    '''
    Wrapper around IOC Multiling XLSX file.
    '''

    _NUM_HEADER_ROWS = 3

    def __init__(self, file_name):
        logging.info(f'Loading {file_name}')
        workbook = openpyxl.load_workbook(file_name, read_only=True)
        self._worksheet = workbook['List']
        self.fields = []
        for row in self._worksheet.iter_rows(
                min_row=1, max_row=Multiling._NUM_HEADER_ROWS, values_only=True):
            if not self.fields:
                self.fields = [None] * len(row)
            for i, cell in enumerate(row):
                if cell and not self.fields[i]:
                    self.fields[i] = cell

    def merged_rows(self):
        '''
        Merges staggered rows into one and yields each row as a dict, except
        the header rows.
        '''
        restart_column = self.fields.index('Scientific Name')
        merged_row = None
        for row in self._worksheet.iter_rows(
                min_row=Multiling._NUM_HEADER_ROWS + 1, values_only=True):
            if row[restart_column]:
                if merged_row and merged_row['Scientific Name']: # pylint: disable=unsubscriptable-object
                    yield merged_row
                merged_row = {field: None for field in self.fields}
            if merged_row:
                for field, value in zip(self.fields, row):
                    if value:
                        merged_row[field] = value
        yield merged_row


class Comparison:
    '''
    Wrapper around "IOC versus other lists" XLSX file.
    '''

    def __init__(self, file_name):
        logging.info(f'Loading {file_name}')
        workbook = openpyxl.load_workbook(file_name, read_only=True)
        self._worksheet = workbook.active
        self.fields = [cell.value for cell in next(self._worksheet.rows)]

    def rows(self):
        '''
        Yields each row as a dict, except header rows.
        '''
        for number, row in enumerate(self._worksheet.iter_rows(min_row=2, values_only=True)):
            d = dict(zip(self.fields, row))
            d['_number'] = number + 2 # 1-based, and we skipped one at the top.
            yield d


def add_args(parser):
    parser.add_argument(
        '--ioc_multiling_file',
        default=os.path.join(os.path.dirname(__file__), 'sources', 'Multiling IOC 9.1b.xlsx'),
        help='Path to the multilingual Excel file downloaded from IOC')
    parser.add_argument(
        '--ioc_comparison_file',
        default=os.path.join(os.path.dirname(__file__), 'sources', 'IOC_9.1_vs_other_lists.xlsx'),
        help='Path to the multilingual Excel file downloaded from IOC')


def main(args, session):
    logging.info('Deleting existing species and common names')
    session.query(CommonName).delete()
    session.query(Species).delete()

    comparison = Comparison(args.ioc_comparison_file)
    logging.info(f'Found column headings: {comparison.fields}')

    ioc_field = [f for f in comparison.fields if 'ioc world bird list' in f.lower()][0]
    clements_field = [f for f in comparison.fields if 'clements checklist of birds of the world' in f.lower()][0]

    ioc_to_species_id = {}
    ioc_to_clements = {}
    for row in comparison.rows():
        ioc = (row[ioc_field] or '').strip()
        clements = (row[clements_field] or '').strip()
        if ioc:
            ioc_to_species_id[ioc] = row['_number']
            if clements:
                ioc_to_clements[ioc] = clements

    multiling = Multiling(args.ioc_multiling_file)
    logging.info(f'Found column headings: {multiling.fields}')

    for field in multiling.fields:
        language_code = _LANGUAGE_MAPPING.get(field)
        if not language_code:
            if not re.match(r'^(No\d.*|Order|Family|Scientific Name)$', field):
                raise ValueError(f'Do not know what to do with column {field}; '
                                 f'maybe it needs to be added as a language?')
            continue

    for row in multiling.merged_rows():
        scientific_name = row['Scientific Name']

        species = session.query(Species)\
            .options(joinedload(Species.common_names))\
            .filter(Species.scientific_name == scientific_name)\
            .one_or_none()
        if species:
            logging.info(f'Already have species {species.scientific_name} '
                         f'(id {species.species_id})')
        else:
            scientific_name = row['Scientific Name']
            species = Species(
                species_id=ioc_to_species_id[scientific_name],
                scientific_name=scientific_name,
                scientific_name_clements=ioc_to_clements.get(scientific_name, None))
            session.add(species)
            session.flush()
            logging.info(f'Added new species {species.scientific_name} (id {species.species_id})')

        for field, value in row.items():
            language_code = _LANGUAGE_MAPPING.get(field)
            if language_code:
                common_names = [
                    common_name for common_name in species.common_names
                    if common_name.language_code == language_code
                ]
                if common_names:
                    common_name = common_names[0]
                else:
                    common_name = CommonName(species_id=species.species_id,
                                             language_code=language_code)
                    session.add(common_name)
                if common_name.common_name != value:
                    common_name.common_name = value
