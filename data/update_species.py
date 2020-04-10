#!/usr/bin/env python

'''
Traverses the official IOC World Bird List
(https://www.worldbirdnames.org/ioc-lists/master-list-2/), the multilingual
version thereof, to produce a machine-readable list of all bird species.

Also assigns stable, unique species ids.
'''

import argparse
import logging
import os
import re
import sys

import openpyxl

from species import Species, SpeciesList


_LANGUAGE_MAPPING = {
    'English': 'en',
    'Afrikaans': 'af',
    'Catalan': 'ca',
    # https://stackoverflow.com/questions/4892372/language-codes-for-simplified-chinese-and-traditional-chinese
    'Chinese': 'zh_CN',
    'Chinese (Traditional)': 'zh_TW',
    'Czech': 'ics',
    'Danish': 'da',
    'Dutch': 'nl',
    'Estonian': 'et',
    'Finnish': 'fi',
    'French': 'fr',
    'German': 'de',
    'Hungarian': 'hu',
    'Icelandic': 'is',
    'Indonesian': 'id',
    'Italian': 'it',
    'Japanese': 'ja',
    'Latvian': 'lv',
    'Lithuanian': 'lt',
    'Northern Sami': 'se',
    'Norwegian': 'no',
    'Polish': 'pl',
    'Portuguese': 'pt',
    'Russian': 'ru',
    'Slovak': 'sk',
    'Slovenian': 'sl',
    'Spanish': 'es',
    'Swedish': 'sv',
    'Thai': 'th',
    'Ukrаiniаn': 'uk',
}


class Multiling:
    '''
    Wrapper around IOC Multiling XLSX file.
    '''

    _NUM_HEADER_ROWS = 3

    def __init__(self, file_name):
        logging.info(f'Loading {file_name}')
        workbook = openpyxl.load_workbook(file_name, read_only=True)
        self._worksheet = workbook['List']
        self.fields = []
        for row in self._worksheet.iter_rows(
                min_row=1, max_row=Multiling._NUM_HEADER_ROWS, values_only=True):
            if not self.fields:
                self.fields = [None] * len(row)
            for i, cell in enumerate(row):
                if cell and not self.fields[i]:
                    self.fields[i] = cell

    def merged_rows(self):
        '''
        Merges staggered rows into one and yields each row as a dict, except
        the header rows.
        '''
        restart_column = self.fields.index('Scientific Name')
        merged_row = None
        for row in self._worksheet.iter_rows(
                min_row=Multiling._NUM_HEADER_ROWS + 1, values_only=True):
            if row[restart_column]:
                if merged_row and merged_row['Scientific Name']: # pylint: disable=unsubscriptable-object
                    yield merged_row
                merged_row = {field: None for field in self.fields}
            if merged_row:
                for field, value in zip(self.fields, row):
                    if value:
                        merged_row[field] = value
        yield merged_row


def _main():
    logging.basicConfig(level=logging.INFO)

    parser = argparse.ArgumentParser()
    parser.add_argument(
        '--ioc_multiling_file',
        default=os.path.join(os.path.dirname(__file__), 'sources', 'Multiling IOC 9.1b.xlsx'),
        help='Path to the multilingual Excel file downloaded from IOC')
    parser.add_argument(
        '--species_list_file',
        default=SpeciesList.DEFAULT_FILE_NAME,
        help='File of species list to be created/updated')
    args = parser.parse_args()

    multiling = Multiling(args.ioc_multiling_file)
    logging.info(f'Found column headings: {multiling.fields}')

    species_list = SpeciesList()
    try:
        species_list.load(args.species_list_file)
        logging.info(f'Read {args.species_list_file} containing {len(species_list)} species')
    except FileNotFoundError:
        logging.info(f'Could not find {args.species_list_file}, starting afresh')

    for field in multiling.fields:
        language_code = _LANGUAGE_MAPPING.get(field)
        if not language_code:
            if not re.match(r'^(No\d.*|Order|Family|Scientific Name)$', field):
                raise ValueError(f'Do not know what to do with column {field}; '
                                 f'maybe it needs to be added as a language?')
            continue

    for row in multiling.merged_rows():
        scientific_name = row['Scientific Name']
        try:
            species = species_list.get_species(scientific_name)
        except KeyError:
            species = Species(scientific_name=row['Scientific Name'], species_id=None)
            species_list.add_species(species)
            logging.info(f'Added new species {species.scientific_name} (id {species.species_id})')
        for field, value in row.items():
            language_code = _LANGUAGE_MAPPING.get(field)
            if language_code:
                species.common_names[language_code] = value

    species_list.save(args.species_list_file)


if __name__ == '__main__':
    sys.exit(_main())
